from selenium import webdriver
from openpyxl import Workbook
from time import sleep
import os.path
import pickle
from datetime import datetime

wb = Workbook()
ws = wb.active

# Fill IDs of stickers each on new line in sticker_ids.txt
ids_of_stickers = [line.strip() for line in open('sticker_ids.txt', 'r', encoding='utf-8').readlines()
                   if line.strip().isdigit()]
names_of_stickers = dict()

names_lines = open('csgo_english.txt', 'r', encoding='utf-8').readlines()

with open('items_game.txt', 'r', encoding='utf-8') as current_items:
    lines = current_items.readlines()
    for i, line in enumerate(lines[::-1]):
        if len(line.strip()) > 2 and line.strip()[1:-1] in ids_of_stickers:
            while "item_name" not in lines[-1 - i]:
                i -= 1
            if '#StickerKit' not in lines[-1 - i]:
                continue
            item_name = lines[-1 - i][:lines[-1 - i].rfind('"')]
            item_name = item_name[item_name.rfind('"') + 2:]
            for line2 in names_lines[::-1]:
                if f'"{item_name}"' in line2:
                    item_name = line2[:line2.rfind('"')]
                    item_name = item_name[item_name.rfind('"') + 1:]
                    names_of_stickers[line.strip()[1:-1]] = item_name
                    break
print(names_of_stickers)

starting_url = 'https://steamcommunity.com/openid/login?openid.claimed_id=http://specs.openid.net/auth/2.0/identifier_select&openid.identity=http://specs.openid.net/auth/2.0/identifier_select&openid.mode=checkid_setup&openid.ns=http://specs.openid.net/auth/2.0&openid.realm=https://csgofloat.com&openid.return_to=https://csgofloat.com/'

options = webdriver.ChromeOptions()
binary_yandex_driver_file = 'chromedriver.exe'

with webdriver.Chrome(binary_yandex_driver_file, options=options) as driver:
    driver.get(starting_url)
    if os.path.isfile('steam_cookies'):
        for cookie in pickle.load(open('steam_cookies', 'rb')):
            driver.add_cookie(cookie)
        driver.refresh()
        if driver.find_elements_by_xpath('//*[@id="steamPassword"]'):
            driver.delete_all_cookies()
            driver.refresh()
    _ = input('Press Enter when authorised')
    driver.get('https://steamcommunity.com/')

    pickle.dump(driver.get_cookies(), open('steam_cookies', 'wb'))

    count_xpath = '/html/body/app-root/div/div[2]/app-float-db/div/div/app-float-dbtable/div/div/mat-card'
    i = 1

    for id in ids_of_stickers:
        address_n = f'A{i}'
        address_q = f'A{i + 1}'
        search_url = 'https://csgofloat.com/db?min=0&max=1&stickers='
        m = []
        for g in range(1, 5):
            driver.get(search_url + str([{'i': id} for _ in range(g)]).replace(' ', '').replace("'", '"'))
            k = 0
            while True:
                if k == 3:
                    k = 0
                    driver.refresh()
                res = driver.find_elements_by_xpath(count_xpath)
                sleep(1.5)
                if res and res[0].text.split()[1].replace(',', '').replace('~','').isdigit():
                    break
                k += 1
            tmp = res[0].text.split()[1].replace(',', '')
            tmp = int(tmp)
            m.append(tmp)
        for g in range(3):
            m[g] = m[g] - m[g + 1]
            m[g] = m[g] * (g + 1)
        m[3] = m[3] * 4
        quant = sum(m)
        ws[address_n] = names_of_stickers[id]
        ws[address_q] = quant
        i += 2

    driver.close()
    wb.save(f"{datetime.now().strftime('%Y-%m-%d')}.xlsx")
